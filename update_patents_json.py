"""
update_patents_json.py

HARCO_patent_list.xlsx 파일을 읽어 특허 실적을 publication/patents.json에 추가합니다.

사용법:
    python update_patents_json.py \
        --xlsx  _data/HARCO_patent_list.xlsx \
        --json  publication/patents.json

옵션:
    --xlsx   엑셀 파일 경로 (기본값: _data/HARCO_patent_list.xlsx)
    --json   JSON 파일 경로 (기본값: publication/patents.json)
    --dry-run  실제 파일을 수정하지 않고 결과만 출력
"""

import argparse
import json
import os
import re
import subprocess
import sys

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    print("[setup] openpyxl 설치 중...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook


# ─────────────────────────────────────────────
# 헬퍼 함수
# ─────────────────────────────────────────────

def extract_year(date_str):
    """'2023.02.22' 또는 '2023-02-22' 형태의 문자열에서 연도를 추출합니다."""
    if not date_str:
        return None
    m = re.match(r"(\d{4})", str(date_str).strip())
    return int(m.group(1)) if m else None


def determine_status(row: dict) -> str:
    """등록번호/출원번호 유무에 따라 특허 상태를 결정합니다."""
    if row.get("등록번호"):
        return "registered"   # 등록
    if row.get("출원번호"):
        return "applied"      # 출원
    return "pending"          # 미출원 (준비 중)


def build_patent_entry(row):
    """
    엑셀의 한 행(dict)을 patents.json 항목으로 변환합니다.

    반환하는 필드:
        title        발명명칭
        inventors    발명자 (콤마 구분 문자열)
        applicant    출원인
        country      국가
        type         구분 (특허/실용신안 등)
        project      과제명
        app_no       출원번호
        app_date     출원일자
        reg_no       등록번호
        reg_date     등록일자
        year         연도 (등록 > 출원 우선)
        status       registered | applied | pending
        category     "patent"  (patents.json 과 구분용)
    """
    reg_year = extract_year(row.get("등록일자"))
    app_year = extract_year(row.get("출원일자"))
    year = reg_year or app_year  # 등록연도 우선, 없으면 출원연도

    return {
        "title":     row.get("발명명칭", "").strip(),
        "inventors": row.get("발명자", ""),
        "applicant": row.get("출원인", ""),
        "country":   row.get("국가", "대한민국"),
        "type":      row.get("구분", "특허"),
        "project":   row.get("과제명") or "",
        "app_no":    row.get("출원번호") or "",
        "app_date":  row.get("출원일자") or "",
        "reg_no":    row.get("등록번호") or "",
        "reg_date":  row.get("등록일자") or "",
        "year":      year,
        "status":    determine_status(row),
        "category":  "patent",
    }


# ─────────────────────────────────────────────
# 엑셀 읽기
# ─────────────────────────────────────────────

COLUMNS = ["관리번호", "국가", "구분", "발명명칭", "발명자", "출원인",
           "과제명", "협약기관", "출원번호", "출원일자", "등록번호", "등록일자"]


def read_xlsx(path):
    """엑셀 파일에서 유효한 특허 행만 읽어 dict 리스트로 반환합니다."""
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    patents = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # 헤더 행 건너뛰기
        if i == 0:
            continue

        # 빈 행 건너뛰기
        if all(v is None for v in row):
            continue

        # 발명명칭이 없으면 건너뛰기
        title = row[3]
        if not title:
            continue

        row_dict = dict(zip(COLUMNS, row))
        patents.append(row_dict)

    wb.close()
    return patents


# ─────────────────────────────────────────────
# JSON 읽기 / 쓰기
# ─────────────────────────────────────────────

def load_json(path):
    if not os.path.exists(path):
        print(f"[INFO] {path} 파일이 없습니다. 새로 생성합니다.")
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save_json(path, data):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


# ─────────────────────────────────────────────
# 중복 제거 로직
# ─────────────────────────────────────────────

def is_duplicate(existing, new_entry):
    """
    이미 존재하는 특허 목록에서 동일 항목이 있는지 확인합니다.
    - 등록번호가 있으면 등록번호로 비교
    - 출원번호가 있으면 출원번호로 비교
    - 둘 다 없으면 title로 비교
    """
    for item in existing:
        if item.get("category") != "patent":
            continue
        reg = new_entry.get("reg_no")
        app = new_entry.get("app_no")
        if reg and item.get("reg_no") == reg:
            return True
        if app and item.get("app_no") == app:
            return True
        if (not reg and not app) and item.get("title") == new_entry.get("title"):
            return True
    return False


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="HARCO 특허 목록을 patents.json에 업데이트합니다.")
    parser.add_argument("--xlsx",    default="_data/HARCO_patent_list.xlsx",
                        help="엑셀 파일 경로")
    parser.add_argument("--json",    default="publication/patents.json",
                        help="JSON 파일 경로")
    parser.add_argument("--dry-run", action="store_true",
                        help="파일을 수정하지 않고 결과만 출력")
    args = parser.parse_args()

    # 1. 엑셀 읽기
    print(f"[1/4] 엑셀 파일 읽는 중: {args.xlsx}")
    raw_patents = read_xlsx(args.xlsx)
    print(f"      → 총 {len(raw_patents)}건 읽음")

    # 2. JSON 읽기
    print(f"[2/4] JSON 파일 읽는 중: {args.json}")
    existing = load_json(args.json)
    print(f"      → 기존 항목 {len(existing)}건")

    # 3. 신규 항목 추출 (중복 제외)
    print("[3/4] 신규 특허 항목 추출 중...")
    new_entries = []
    skipped = []
    for row in raw_patents:
        entry = build_patent_entry(row)
        if is_duplicate(existing, entry):
            skipped.append(entry["title"])
        else:
            new_entries.append(entry)

    print(f"      → 신규 {len(new_entries)}건 / 중복 건너뜀 {len(skipped)}건")
    if skipped:
        for t in skipped:
            print(f"        (skip) {t}")

    # 4. 저장 또는 출력
    if not new_entries:
        print("[4/4] 추가할 신규 항목이 없습니다.")
        return

    # 특허 항목을 기존 목록 앞에 추가 (연도 내림차순 정렬)
    merged = new_entries + existing
    merged.sort(key=lambda x: (x.get("year") or 0), reverse=True)

    if args.dry_run:
        print("[4/4] --dry-run 모드: 아래 내용이 추가될 예정입니다.\n")
        print(json.dumps(new_entries, ensure_ascii=False, indent=4))
    else:
        print(f"[4/4] JSON 파일 저장 중: {args.json}")
        save_json(args.json, merged)
        print(f"      → 저장 완료! 총 {len(merged)}건 (신규 {len(new_entries)}건 추가)")


if __name__ == "__main__":
    main()